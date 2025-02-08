import socket
from time import sleep
import openpyxl
from datetime import datetime


def user_date(user_number):
    wb = openpyxl.load_workbook("table.xlsx")
    sheets = {str(i).split('"')[1]: i for i in wb.worksheets}
    users = sheets.get('users')

    if not users:
        return False

    users_list = []
    cur = 1
    spaces = 0

    users_id_list = list(map(lambda x: x[0].value, list(users.iter_rows())))
    dates_list = list(map(lambda x: x[2].value, list(users.iter_rows())))
    print(users_id_list, dates_list)
    return dates_list[users_id_list.index(user_number)]


def check_user(user_number):
    wb = openpyxl.load_workbook("table.xlsx")
    sheets = {str(i).split('"')[1]: i for i in wb.worksheets}
    users = sheets.get('users')

    if not users:
        return False

    users_list = []
    cur = 1
    spaces = 0

    users_list = list(map(lambda x: x[0].value, list(users.iter_rows())))
    print(users_list)

    return user_number in users_list


def check_info(data):
    words = ['enter', 'exit']
    hex_alp = '0123456789abcdef'

    if not (data.startswith('(') and data.endswith(')')):
        return False

    data = data[1:-1].split()

    if len(data) != 2 or data[1] not in words:
        return False

    hex_part = data[0]
    if not (hex_part.startswith('0x') and all(c in hex_alp for c in hex_part[2:])):
        return False

    return True


HOST = "192.168.0.101"
PORT = 1337

with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
    s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)  # Разрешаем повторное использование порта
    s.bind((HOST, PORT))
    s.listen()

    while True:
        try:
            conn, addr = s.accept()
            conn.settimeout(10)  # Тайм-аут на прием данных
            timestamp = datetime.now().strftime("%d-%m-%Y %H:%M:%S")  # Получаем текущее время
            print(f"[{timestamp}] Connected by {addr}")

            with conn:
                while True:
                    try:
                        data = conn.recv(1024)
                        if not data:
                            break
                        data = data.decode('utf-8').strip().lower()
                        response = b'FALSE'

                        if check_info(data):
                            with open('info.txt', 'a') as f:
                                f.write(f"[{timestamp}] {data}\n")
                            user_id = data[1:-1].split()[0]
                            if (check_user(user_id) and not datetime.now() > user_date(user_id)) or 'exit' in data.lower():
                                response = b'TRUE'
                        else:
                            with open('dump.txt', 'a') as f:
                                f.write(f"[{timestamp}] {data}\n")

                        conn.sendall(response)
                    except socket.timeout:
                        print(f"[{timestamp}] Timeout: клиент не отправил данные")
                        break
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Ошибка: {e}")
