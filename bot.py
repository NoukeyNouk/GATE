import telebot
import openpyxl
from openpyxl.styles import numbers
from datetime import datetime
from secrets import secret_token


def correct_format(info):
    if info.count(';') != 2:
        return False
    info = tuple(info.split(';'))

    if not all(list(map(lambda x: x in filter_flag, list(info[0].lower())))):
        return False

    try:
        datetime.strptime(info[2], '%d.%m.%Y')
    except ValueError:
        return False

    if not ''.join(info[1].split()).isalpha():
        return False

    return True


def edit_user(info, remove=False):
    wb = openpyxl.load_workbook("table.xlsx")
    sheets = {str(i).split('"')[1]: i for i in wb.worksheets}
    users = sheets.get('users')

    if not users:
        return 'Not found.'

    users_list = []
    cur = 1
    spaces = 0

    users_list = [list(map(lambda x: x[0].value, list(users.iter_rows()))),
                  list(map(lambda x: x[1].value, list(users.iter_rows()))),
                  list(map(lambda x: x[2].value, list(users.iter_rows())))]

    full_list = []

    for i in range(len(users_list[0])):
        if users_list[0][i]:
            full_list.append((users_list[0][i], users_list[1][i], users_list[2][i]))

    print(full_list)

    if not remove:
        if info in full_list:
            return 'Already in database.'

        full_list.append(info)

    else:
        if info not in full_list:
            return 'Not in list'

        for row in users.iter_rows():
            for cell in row:
                cell.value = None

        full_list.remove(info)

    for i in range(len(full_list)):
        for j in range(3):
            users.cell(row=i + 1, column=j + 1).value = full_list[i][j]
            if j == 2:
                users.cell(row=i + 1, column=j + 1).number_format = 'DD.MM.YYYY'

    wb.save("table.xlsx")
    return 'Saved.'


admin_list = [847893338, 2119196274]
token = secret_token
bot = telebot.TeleBot(token)
save_flag = False
remove_flag = False
filter_flag = '0123456789abcdefx'


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, 'Hello')

@bot.message_handler(commands=['add'])
def add_message(message):
    global save_flag
    bot.send_message(message.chat.id, 'Type new user.')
    save_flag = True

@bot.message_handler(commands=['remove'])
def remove_message(message):
    global remove_flag
    bot.send_message(message.chat.id, 'Type user.')
    remove_flag = True

@bot.message_handler(commands=['send_logs'])
def send_logs_message(message):
    if message.from_user.id in admin_list:
        with open("info.txt", "rb") as file:
            bot.send_document(message.chat.id, file)
        with open("dump.txt", "rb") as file:
            bot.send_document(message.chat.id, file)
        with open("table.xlsx", "rb") as file:
            bot.send_document(message.chat.id, file)
    else:
        bot.send_message(message.chat.id, 'Let it be.')


@bot.message_handler(content_types=['text'])
def text_message(message):
    global save_flag
    global remove_flag
    if save_flag and message.from_user.id in admin_list:
        info = message.text
        if not correct_format(info):
            bot.send_message(message.chat.id, 'Incorrect format')
        else:
            info = info.split(';')
            info[2] = datetime.strptime(info[2], '%d.%m.%Y')
            info = tuple(info)

            try:
                response = edit_user(info)
                bot.send_message(message.chat.id, response)
            except Exception as e:
                bot.send_message(message.chat.id, f'Error: {e}')
        save_flag = False
    elif remove_flag and message.from_user.id in admin_list:
        info = message.text
        if not correct_format(info):
            bot.send_message(message.chat.id, 'Incorrect format')
        else:
            info = info.split(';')
            info[2] = datetime.strptime(info[2], '%d.%m.%Y')
            info = tuple(info)

            try:
                response = edit_user(info, remove=True)
                bot.send_message(message.chat.id, response)
            except Exception as e:
                bot.send_message(message.chat.id, f'Error: {e}')
        remove_flag = False
    else:
        bot.send_message(message.chat.id, 'Let it be.')


bot.infinity_polling()
